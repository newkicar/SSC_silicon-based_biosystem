"""
加班基础数据自动构建脚本

从SAP考勤API获取全年考勤数据，结合花名册计算：
- 每人每月加班时长（白领/蓝领不同算法）
- 部门/中心/公司月加班统计
- 出勤率（个人/部门/中心/公司）

输出: databases/加班基础数据.xlsx（追加写入，当月覆盖，跨月累加）
"""

import sys
import os
import io
import base64
import json
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import pandas as pd
import requests
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_v1_5

# ==================== 配置 ====================
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_DIR = PROJECT_ROOT / "databases"
ROSTER_PATH = DB_DIR / "员工花名册.xlsx"
EXCLUDE_PATH = DB_DIR / "assistance" / "不统计的人.xlsx"
OUTPUT_PATH = DB_DIR / "加班基础数据.xlsx"

SAP_API_URL = "http://{{SAP_API_HOST}}:8080/system/sap/queryAttendance"
SAP_AUTH_CODE = "{{SAP_AUTH_CODE_ATTENDANCE}}"
SAP_PUBLIC_KEY = "{{RSA_PUBLIC_KEY_ATTENDANCE}}"


def get_default_date_range():
    """
    默认日期范围：
    - 当天=当月1日 → 拉取上个月整月
    - 当天≥2日 → 拉取本月1日~昨天
    """
    from calendar import monthrange

    today = datetime.now()
    year = today.year
    month = today.month
    day = today.day

    if day == 1:
        # 上个月整月
        if month == 1:
            prev_year, prev_month = year - 1, 12
        else:
            prev_year, prev_month = year, month - 1
        _, last_day = monthrange(prev_year, prev_month)
        begin = f"{prev_year}-{prev_month:02d}-01"
        end = f"{prev_year}-{prev_month:02d}-{last_day:02d}"
    else:
        # 本月1日 ~ 昨天
        yesterday = today - timedelta(days=1)
        begin = f"{year}-{month:02d}-01"
        end = yesterday.strftime("%Y-%m-%d")
    return begin, end


def generate_month_ranges(begin_date_str, end_date_str):
    """
    根据起止日期生成按月拆分的 [(month_num, begin, end), ...] 列表。
    例如 "2026-06-01" ~ "2026-06-17" → [(6, "2026-06-01", "2026-06-17")]
         "2026-01-01" ~ "2026-06-17" → [(1, "2026-01-01", "2026-01-31"), ..., (6, "2026-06-01", "2026-06-17")]
    """
    from calendar import monthrange

    begin_dt = datetime.strptime(begin_date_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")

    ranges = []
    cur = begin_dt
    while cur <= end_dt:
        year = cur.year
        month = cur.month
        # 当月最后一天
        _, last_day = monthrange(year, month)
        month_end = datetime(year, month, last_day)

        # 取 min(月末, end_dt)
        actual_end = min(month_end, end_dt)
        # 取 max(月初, begin_dt)
        actual_begin = max(cur, datetime(year, month, 1))

        ranges.append(
            (month, actual_begin.strftime("%Y-%m-%d"), actual_end.strftime("%Y-%m-%d"))
        )

        # 下个月1日
        if month == 12:
            cur = datetime(year + 1, 1, 1)
        else:
            cur = datetime(year, month + 1, 1)

    return ranges


# 排除的leaveType
EXCLUDE_LEAVE_TYPES = {"产假", "陪产假", "流产假", "工伤假", "婚假"}


# ==================== RSA加密 ====================
def rsa_encrypt(plain_text: str) -> str:
    key_der = base64.b64decode(SAP_PUBLIC_KEY)
    rsa_key = RSA.import_key(key_der)
    cipher = PKCS1_v1_5.new(rsa_key)
    encrypted = cipher.encrypt(plain_text.encode("utf-8"))
    return base64.b64encode(encrypted).decode("utf-8")


def generate_auth_info():
    today = datetime.now().strftime("%Y%m%d")
    return rsa_encrypt(f"{today}&{SAP_AUTH_CODE}")


# ==================== 工号标准化 ====================
def _normalize_eid_suffix(raw_eid: str) -> str:
    """提取工号后6位纯数字。
    花名册工号可能带L前缀(L11095)、00前缀(0011095)、或原样(11095)，
    [ERP系统接口]工号统一00前缀(0011095=8位)。后6位一致，以此为内部统一key。
    """
    s = str(raw_eid).strip()
    digits = "".join(c for c in s if c.isdigit())
    return digits[-6:] if len(digits) >= 6 else digits


# ==================== Step 1: 读取花名册 ====================
def load_roster():
    """读取花名册，返回 {工号后6位: {字段...}} 字典"""
    print("[1] 读取花名册...")
    wb = openpyxl.load_workbook(str(ROSTER_PATH), read_only=True, data_only=True)
    ws = wb["花名册"]

    headers = []
    roster = {}
    skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        d = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        raw_eid = str(d.get("员工号", "")).strip()
        if not raw_eid or raw_eid == "None":
            continue
        eid_suffix = _normalize_eid_suffix(raw_eid)
        if len(eid_suffix) < 6:
            skipped += 1
            continue
        roster[eid_suffix] = d

    wb.close()
    print(f"  花名册: {len(roster)} 人 (跳过{skipped}条无效工号)")
    return roster


# ==================== Step 2: 读取排除名单 ====================
def load_exclusion_list():
    """读取不统计的人.xlsx"""
    print("[2] 读取排除名单...")
    if not EXCLUDE_PATH.exists():
        print(f"  文件不存在: {EXCLUDE_PATH}，跳过")
        return []

    wb = openpyxl.load_workbook(str(EXCLUDE_PATH), read_only=True, data_only=True)
    ws = wb.active

    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 表头
            continue
        if len(row) < 5:
            continue
        eid = str(row[1]).strip() if row[1] else ""
        start_date = str(row[3]).strip() if row[3] else ""
        end_date = str(row[4]).strip() if row[4] else ""
        if eid and start_date and end_date:
            eid_suffix = _normalize_eid_suffix(eid)
            records.append(
                {
                    "eid_suffix": eid_suffix,
                    "start": start_date,
                    "end": end_date,
                }
            )
    wb.close()
    print(f"  排除名单: {len(records)} 条")
    return records


def is_excluded_by_list(eid_suffix, attendance_date_str, exclusion_list):
    """检查是否在排除名单中"""
    for exc in exclusion_list:
        if eid_suffix == exc["eid_suffix"]:
            if exc["start"] <= attendance_date_str <= exc["end"]:
                return True
    return False


# ==================== Step 3: 筛选需要查询的员工 ====================
def get_eligible_employees(roster, target_month=None):
    """
    筛选需要查询考勤的员工：
    - 大级别1-8
    - 非离职（无离职日期或离职日期为空）
    - 蓝领白领非"实习生"
    - 部门不包含"大客户管理部"
    - 如果target_month: 排除当月入职/离职的
    """
    eligible = {}
    for eid, info in roster.items():
        # 大级别
        level = info.get("大级别")
        try:
            level_int = int(float(str(level)))
        except (ValueError, TypeError):
            continue
        if level_int < 1 or level_int > 8:
            continue

        # 实习生
        collar = str(info.get("蓝领白领", "")).strip()
        if collar == "实习生":
            continue

        # 大客户管理部
        dept = str(info.get("部门", "")).strip()
        if "大客户管理部" in dept:
            continue

        # 当月入职/离职排除
        if target_month:
            entry_date = info.get("入职时间")
            leave_date = info.get("离职日期")

            # 入职日期年月匹配 → 排除
            if entry_date:
                entry_str = str(entry_date).strip()
                try:
                    if " " in entry_str:
                        entry_str = entry_str.split(" ")[0]
                    entry_dt = datetime.strptime(entry_str[:10], "%Y-%m-%d")
                    if entry_dt.year == 2026 and entry_dt.month == target_month:
                        continue
                except (ValueError, TypeError):
                    pass

            # 离职日期年月匹配 → 排除
            if leave_date:
                leave_str = str(leave_date).strip()
                if leave_str not in ("", "None", "NaT"):
                    try:
                        if " " in leave_str:
                            leave_str = leave_str.split(" ")[0]
                        leave_dt = datetime.strptime(leave_str[:10], "%Y-%m-%d")
                        if leave_dt.year == 2026 and leave_dt.month == target_month:
                            continue
                    except (ValueError, TypeError):
                        pass

        eligible[eid] = info

    return eligible


# ==================== Step 4: [ERP系统接口]获取考勤 ====================
def fetch_attendance_batch(employees, begin_date, end_date, auth_info):
    """批量获取员工考勤数据"""
    all_records = []
    success = 0
    fail = 0

    for idx, (eid_suffix, info) in enumerate(employees.items()):
        params = {
            "personNo": "00" + eid_suffix,  # SAP统一 00 + 后6位
            "beginDate": begin_date,
            "endDate": end_date,
            "authInfo": auth_info,
        }
        try:
            resp = requests.post(
                SAP_API_URL,
                json=params,
                headers={"Content-Type": "application/json"},
                timeout=10,
            )
            resp.raise_for_status()
            result = resp.json()
            if result.get("code") == 200:
                data = result.get("data", [])
                if isinstance(data, list):
                    for rec in data:
                        if isinstance(rec, dict):
                            rec["__员工号"] = eid_suffix  # 统一用后6位
                            all_records.append(rec)
                    success += 1
            else:
                fail += 1
        except Exception:
            fail += 1

        if (idx + 1) % 100 == 0:
            print(f"    进度: {idx + 1}/{len(employees)} (成功{success}, 失败{fail})")
        time.sleep(0.05)

    print(f"    完成: 成功{success}, 失败{fail}, 总记录{len(all_records)}")
    return all_records


# ==================== Step 5: 解析leaveDuration ====================
def parse_leave_duration(leave_duration_str):
    """解析逗号分隔的leaveDuration，返回float列表"""
    if not leave_duration_str or str(leave_duration_str).strip() in ("None", "", "nan"):
        return []
    s = str(leave_duration_str).strip()
    parts = []
    for p in s.split(","):
        p = p.strip()
        try:
            parts.append(float(p))
        except ValueError:
            parts.append(0.0)
    return parts


def parse_leave_types(leave_type_str):
    """解析逗号分隔的leaveType，返回str列表"""
    if not leave_type_str or str(leave_type_str).strip() in ("None", "", "nan"):
        return []
    return [t.strip() for t in str(leave_type_str).split(",") if t.strip()]


# ==================== Step 6: 计算单日加班时长 ====================
def compute_daily_overtime(record, collar):
    """
    计算单日加班时长
    编制类型含"Salary"（含Salary/Indirect Salary）: attendanceDuration + leaveDuration(合计) - 8
    否则（纯蓝领）: overtimeDuration + overtimeHDuration + overtimeWDuration
    """
    att_duration = float(record.get("attendanceDuration", 0) or 0)

    if "Salary" in str(collar):
        leave_dur_list = parse_leave_duration(record.get("leaveDuration"))
        leave_total = sum(leave_dur_list)
        overtime = att_duration + leave_total - 8
        return max(overtime, 0)  # 不允许负加班
    else:
        # 蓝领
        ot_dur = float(record.get("overtimeDuration", 0) or 0)
        ot_h = float(record.get("overtimeHDuration", 0) or 0)
        ot_w = float(record.get("overtimeWDuration", 0) or 0)
        return ot_dur + ot_h + ot_w


# ==================== Step 7: 计算单日出勤率 ====================
def compute_daily_attendance_rate(record):
    """
    计算个人当日出勤率
    前提: attendanceClasses不包含"假日"和"休息日"
    公式: min((attendanceDuration + 放工 + 出差 + 外勤) / (8 + 旷工), 1)
    """
    classes = str(record.get("attendanceClasses", "")).strip()
    if "假日" in classes or "休息日" in classes:
        return None  # 不参与计算

    att_duration = float(record.get("attendanceDuration", 0) or 0)
    absentee = float(record.get("absenteeDuration", 0) or 0)

    # 解析leaveType和leaveDuration
    leave_types = parse_leave_types(record.get("leaveType"))
    leave_durs = parse_leave_duration(record.get("leaveDuration"))

    # 匹配放工/出差/外勤的时长
    fanggong_hours = 0.0
    chuchai_hours = 0.0
    waiqin_hours = 0.0

    for i, lt in enumerate(leave_types):
        dur = leave_durs[i] if i < len(leave_durs) else 0.0
        if "放工" in lt:
            fanggong_hours += dur
        elif "出差" in lt:
            chuchai_hours += dur
        elif "外勤" in lt:
            waiqin_hours += dur

    numerator = att_duration + fanggong_hours + chuchai_hours + waiqin_hours
    denominator = 8 + absentee

    if denominator <= 0:
        return 0.0

    rate = min(numerator / denominator, 1.0)
    return rate


# ==================== Step 8: 计算夜班/迟到/早退/补卡/漏打卡 ====================
def parse_int_field(val, default=0):
    """安全解析整数字段"""
    try:
        return int(float(str(val or default)))
    except (ValueError, TypeError):
        return default


# ==================== 主流程 ====================
def main():
    import argparse

    parser = argparse.ArgumentParser(description="加班基础数据自动构建")
    parser.add_argument(
        "--begin", type=str, default="", help="起始日期 (YYYY-MM-DD)，默认本月1日"
    )
    parser.add_argument(
        "--end", type=str, default="", help="结束日期 (YYYY-MM-DD)，默认昨天"
    )
    args = parser.parse_args()

    if args.begin and args.end:
        begin_date_str = args.begin
        end_date_str = args.end
    else:
        begin_date_str, end_date_str = get_default_date_range()

    month_ranges = generate_month_ranges(begin_date_str, end_date_str)

    start_time = time.time()
    print("=" * 60)
    print("加班基础数据自动构建")
    print(f"日期范围: {begin_date_str} ~ {end_date_str}")
    print(f"输出: {OUTPUT_PATH}")
    print("=" * 60)

    # 1. 读取花名册
    roster = load_roster()

    # 2. 读取排除名单
    exclusion_list = load_exclusion_list()

    # 3. 准备输出数据
    output_rows = []

    auth_info = generate_auth_info()

    # 4. 按月循环
    for month_num, begin_date, end_date in month_ranges:
        print(f"\n{'='*60}")
        print(f"处理 {month_num} 月: {begin_date} ~ {end_date}")
        print(f"{'='*60}")

        # 筛选当月员工
        employees = get_eligible_employees(roster, target_month=month_num)
        print(f"  当月应统计员工: {len(employees)} 人")

        # 获取考勤数据
        print(f"  调用[ERP系统接口]...")
        records = fetch_attendance_batch(employees, begin_date, end_date, auth_info)

        if not records:
            print(f"  无考勤数据，跳过")
            continue

        # 5. 按员工聚合
        # 每人每天的记录 → 聚合为每人每月
        emp_monthly = defaultdict(
            lambda: {
                "days": [],
                "total_overtime": 0.0,
                "attendance_rates": [],
                "total_attendance_duration": 0.0,
                "night_shift": 0,
                "late": 0,
                "early": 0,
                "absentee": 0.0,
                "workday_ot": 0.0,
                "weekend_ot": 0.0,
                "holiday_ot": 0.0,
                "punch_card_count": 0,  # 补卡=迟到次数
                "miss_card_count": 0,  # 漏打卡
            }
        )

        for rec in records:
            eid = rec.get("__员工号", "")
            att_date = str(rec.get("attendanceDate", "")).strip()
            if not eid or not att_date:
                continue

            # 检查是否在排除名单中
            if is_excluded_by_list(eid, att_date, exclusion_list):
                continue

            info = employees.get(eid, roster.get(eid, {}))
            collar = str(info.get("人员编制", "")).strip()
            if collar == "" or "实习" in collar:
                continue

            # 检查leaveType是否排除
            leave_types = parse_leave_types(rec.get("leaveType"))
            skip = False
            for lt in leave_types:
                if any(ex in lt for ex in EXCLUDE_LEAVE_TYPES):
                    skip = True
                    break
            if skip:
                continue

            em = emp_monthly[eid]
            em["days"].append(att_date)

            # 加班时长
            daily_ot = compute_daily_overtime(rec, collar)
            em["total_overtime"] += daily_ot

            # 出勤率
            daily_rate = compute_daily_attendance_rate(rec)
            if daily_rate is not None:
                em["attendance_rates"].append(daily_rate)

            # 出勤时长
            em["total_attendance_duration"] += float(
                rec.get("attendanceDuration", 0) or 0
            )

            # 迟到/早退/旷工
            em["late"] += parse_int_field(rec.get("lateCount"))
            em["early"] += parse_int_field(rec.get("earlyCount"))
            em["absentee"] += float(rec.get("absenteeDuration", 0) or 0)

            # 补卡次数（迟到次数作为补卡次数）
            if parse_int_field(rec.get("lateCount")) > 0:
                em["punch_card_count"] += parse_int_field(rec.get("lateCount"))

            # 夜班
            classes = str(rec.get("attendanceClasses", ""))
            if "夜班" in classes:
                em["night_shift"] += 1

            # 工作日/双休/法定加班
            em["workday_ot"] += float(rec.get("overtimeWDuration", 0) or 0)
            em["weekend_ot"] += float(rec.get("overtimeHDuration", 0) or 0)
            em["holiday_ot"] += float(rec.get("overtimeADuration", 0) or 0)

        # 6. 生成输出行
        for eid, em in emp_monthly.items():
            info = employees.get(eid, roster.get(eid, {}))
            if not info:
                continue

            # 月加班时长
            monthly_ot = em["total_overtime"]

            # 月出勤率
            rates = em["attendance_rates"]
            personal_attendance = sum(rates) / len(rates) if rates else None

            num_days = len(em["days"])

            # 日均加班
            daily_avg = monthly_ot / num_days if num_days > 0 else 0

            # 日均小于半小时/一小时
            less_half = "是" if 0 < daily_avg < 0.5 else "否"
            less_one = "是" if 0 < daily_avg < 1.0 else "否"

            # 近两个月加班时长大于60（暂标记否，后续可跨月计算）
            over_60 = "是" if monthly_ot > 60 else "否"

            # 加班时长小于1小时的日均时长
            less_1h_daily = daily_avg if daily_avg < 1 else 0

            row = {
                "地区": str(info.get("地区", "")).strip(),
                "公司": str(info.get("公司", "")).strip(),
                "中心": str(info.get("中心", "")).strip(),
                "部门": str(info.get("部门", "")).strip(),
                "二级部门": str(info.get("二级部门", "")).strip(),
                "三级部门": str(info.get("三级部门", "")).strip(),
                "编制类型": str(info.get("人员编制", "")).strip(),
                "蓝领白领": str(info.get("蓝领白领", "")).strip(),
                "员工职级": str(info.get("员工职级", "")).strip(),
                "员工编号": eid,
                "姓名": str(info.get("姓名", "")).strip(),
                "工号姓名": f"{eid} {info.get('姓名', '')}",
                "岗级": str(info.get("岗级", "")).strip(),
                "大级别": info.get("大级别", ""),
                "白领蓝领": str(info.get("蓝领白领", "")).strip(),
                "岗位": str(info.get("岗位", "")).strip(),
                "入职日期": str(info.get("入职时间", "")).strip(),
                "考勤日期": f"2026-{month_num:02d}-01",
                "考勤年份": "2026年",
                "白领员工月人均加班时长": None,  # 后续填充
                "蓝领员工月人均加班时长": None,  # 后续填充
                "部门月加班人数占比": None,  # 后续填充
                "中心月加班人数占比": None,  # 后续填充
                "部门月总人数": None,  # 后续填充
                "中心月总人数": None,  # 后续填充
                "公司月总人数": None,  # 后续填充
                "白领蓝领月总人数": None,  # 后续填充
                "日均小于半小时": less_half,
                "日均小于一小时": less_one,
                "近两个月加班时长大于60小时": over_60,
                "考勤月份": month_num,
                "补卡次数": em["punch_card_count"],
                "漏打卡次数": em["miss_card_count"],
                "早退次数": em["early"],
                "迟到次数": em["late"],
                "迟到和早退": em["late"] + em["early"],
                "夜班次数": em["night_shift"],
                "工作日加班": round(em["workday_ot"], 2),
                "双休加班": round(em["weekend_ot"], 2),
                "法定加班": round(em["holiday_ot"], 2),
                "旷工小时数": round(em["absentee"], 2),
                "当日加班时长": round(monthly_ot, 2),
                "加班时长小于1小时的日均时长": round(less_1h_daily, 2),
                "员工当日出勤率": (
                    round(personal_attendance, 6)
                    if personal_attendance is not None
                    else None
                ),
                "部门当日出勤率": None,  # 后续填充
            }
            output_rows.append(row)

    # 7. 填充聚合字段
    print(f"\n{'='*60}")
    print("填充聚合字段...")
    print(f"{'='*60}")

    # 构建临时DataFrame方便聚合
    df = pd.DataFrame(output_rows)
    if df.empty:
        print("  无数据，退出")
        return

    # 将"后续填充"的列统一转为 float，避免后续写入数值时报 dtype 'str' 错误
    fill_cols = [
        "白领员工月人均加班时长",
        "蓝领员工月人均加班时长",
        "部门月加班人数占比",
        "中心月加班人数占比",
        "部门月总人数",
        "中心月总人数",
        "公司月总人数",
        "白领蓝领月总人数",
        "部门当日出勤率",
    ]
    for col in fill_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["当日加班时长_num"] = pd.to_numeric(df["当日加班时长"], errors="coerce").fillna(
        0
    )
    df["员工当日出勤率_num"] = pd.to_numeric(df["员工当日出勤率"], errors="coerce")

    # 按公司+中心+部门+月份+蓝领白领聚合
    for _, group in df.groupby(["公司", "中心", "部门", "考勤月份"]):
        dept_attendance = group["员工当日出勤率_num"].dropna()
        dept_att_mean = dept_attendance.mean() if len(dept_attendance) > 0 else None

        dept_size = len(group)
        dept_ot_people = (group["当日加班时长_num"] > 0).sum()
        dept_ot_ratio = dept_ot_people / dept_size if dept_size > 0 else 0

        # 中心人数
        center = group.iloc[0]["中心"]
        company = group.iloc[0]["公司"]
        month = group.iloc[0]["考勤月份"]
        center_group = df[(df["中心"] == center) & (df["考勤月份"] == month)]
        company_group = df[(df["公司"] == company) & (df["考勤月份"] == month)]

        center_size = len(center_group)
        company_size = len(company_group)
        center_ot_people = (center_group["当日加班时长_num"] > 0).sum()
        center_ot_ratio = center_ot_people / center_size if center_size > 0 else 0

        # 蓝领白领
        white_collar = group[group["蓝领白领"] == "白领"]
        blue_collar = group[group["蓝领白领"] != "白领"]

        white_avg = (
            white_collar["当日加班时长_num"].mean() if len(white_collar) > 0 else None
        )
        blue_avg = (
            blue_collar["当日加班时长_num"].mean() if len(blue_collar) > 0 else None
        )

        # 填回
        for idx in group.index:
            df.at[idx, "部门当日出勤率"] = (
                round(float(dept_att_mean), 6) if pd.notna(dept_att_mean) else None
            )
            df.at[idx, "部门月加班人数占比"] = round(dept_ot_ratio, 4)
            df.at[idx, "中心月加班人数占比"] = round(center_ot_ratio, 4)
            df.at[idx, "部门月总人数"] = dept_size
            df.at[idx, "中心月总人数"] = center_size
            df.at[idx, "公司月总人数"] = company_size
            df.at[idx, "白领蓝领月总人数"] = len(group)
            df.at[idx, "白领员工月人均加班时长"] = (
                round(float(white_avg), 2) if pd.notna(white_avg) else None
            )
            df.at[idx, "蓝领员工月人均加班时长"] = (
                round(float(blue_avg), 2) if pd.notna(blue_avg) else None
            )

    # 删除临时列
    df = df.drop(columns=["当日加班时长_num", "员工当日出勤率_num"], errors="ignore")

    # 8. 合并写入：删旧月数据 → 追加新月数据
    print(f"\n写入 {OUTPUT_PATH}...")

    # 确定本次涉及的月份
    target_months = set(df["考勤月份"].unique())
    print(f"  本次目标月份: {sorted(target_months)}")

    if OUTPUT_PATH.exists():
        # 读取已有文件
        existing_df = pd.read_excel(str(OUTPUT_PATH))
        print(
            f"  已有文件: {len(existing_df)} 行, 月份 {sorted(existing_df['考勤月份'].unique())}"
        )

        # 删除目标月份的旧数据
        keep_mask = ~existing_df["考勤月份"].isin(target_months)
        kept_df = existing_df[keep_mask]
        removed_count = len(existing_df) - len(kept_df)
        print(f"  删除旧月份数据: {removed_count} 行, 保留 {len(kept_df)} 行")

        # 合并：保留的旧数据 + 新数据
        final_df = pd.concat([kept_df, df], ignore_index=True)
        print(f"  合并后: {len(final_df)} 行")
    else:
        final_df = df
        print(f"  新建文件")

    final_df.to_excel(str(OUTPUT_PATH), index=False, sheet_name="Sheet1")
    print(f"  已保存: {len(final_df)} 行")

    elapsed = time.time() - start_time
    print(f"\n{'='*60}")
    print(f"完成！耗时: {elapsed/60:.1f} 分钟")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
