"""
花名册自动刷新脚本
从清册中提取数据，结合[ERP系统接口]的岗级信息，生成完整的花名册。
"""

import sys
import os
import io
import base64
import json
import glob
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import msoffcrypto
import requests
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_v1_5

# ==================== 配置 ====================
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_DIR = PROJECT_ROOT / "databases"


# 通配解析：自动匹配目录中最新的  清册文件
def _resolve_latest(pattern: str, label: str) -> Path:
    """从 pattern（含 * 通配符）中解析出匹配的最新文件路径。"""
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(
            f"未找到匹配的文件: {pattern}\n"
            f"请确认 H 盘已挂载且目录中存在 '{label}' 清册文件。"
        )
    return Path(matches[-1])  # 取最新（按文件名排序的最后一条）


CC_PATH = _resolve_latest(
    r"{{INTERNAL_HR_PATH}}03-SSC share\05-人员基数数据\新版CC人员清册*.xlsx", "CC"
)
DL_PATH = _resolve_latest(
    r"{{INTERNAL_HR_PATH}}03-SSC share\05-人员基数数据\新版DL人员清册*.xlsx", "DL"
)
print(f"  CC文件: {CC_PATH.name}")
print(f"  DL文件: {DL_PATH.name}")

EXISTING_PATH = DB_DIR / "员工花名册.xlsx"

CC_PASSWORD = "{{EXCEL_PASSWORD_CC}}"
DL_PASSWORD = "{{EXCEL_PASSWORD_CC}}111"

# [ERP系统接口]
SAP_API_URL = "http://{{SAP_API_HOST}}:8080/system/sap/querySapRoster"
SAP_AUTH_CODE = "{{SAP_AUTH_CODE_ROSTER}}"
SAP_PUBLIC_KEY = "{{RSA_PUBLIC_KEY_ROSTER}}"


# ==================== RSA加密工具 ====================
def _rsa_encrypt(plain_text: str, public_key_str: str) -> str:
    try:
        key_der = base64.b64decode(public_key_str)
        rsa_key = RSA.import_key(key_der)
        cipher = PKCS1_v1_5.new(rsa_key)
        encrypted = cipher.encrypt(plain_text.encode("utf-8"))
        return base64.b64encode(encrypted).decode("utf-8")
    except Exception as e:
        print(f"  [错误] RSA加密失败: {e}")
        return ""


def generate_auth_info():
    today = datetime.now().strftime("%Y%m%d")
    plaintext = f"{today}&{SAP_AUTH_CODE}"
    return _rsa_encrypt(plaintext, SAP_PUBLIC_KEY)


# ==================== Step 1: 解密 + 合并 ====================
def decrypt_file(encrypted_path, password):
    """解密OLE2加密的Excel文件，返回BytesIO"""
    with open(encrypted_path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        dec = io.BytesIO()
        file.decrypt(dec)
        dec.seek(0)
        return dec


def read_roster_sheet_from_bytes(file_bytes, sheet_name, source_label):
    """
    从BytesIO中读取指定sheet。
    行1=空行跳过，行2=少量元数据，行3=中文标题保留作为列名，行4=英文标题跳过，行5起=数据。
    返回 (headers, rows) — headers是行3的中文标题，rows是行5起的数据行列表。
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb[sheet_name]

    # 行3 = 中文标题
    headers = []
    seen = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        name = str(val).strip() if val is not None else f"col_{col}"
        # 去重处理（如"学校名称"出现两次）
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    # 行5起 = 数据
    data_rows = []
    for row_idx in range(5, ws.max_row + 1):
        row_vals = []
        for col in range(1, ws.max_column + 1):
            row_vals.append(ws.cell(row=row_idx, column=col).value)
        if any(v is not None for v in row_vals):
            data_rows.append(row_vals)

    wb.close()
    return headers, data_rows


def step1_stack_sheets():
    """Step 1: 解密，合并4个sheet"""
    print("=" * 60)
    print("Step 1: 解密，合并 Employee list + 离职")
    print("=" * 60)

    sources = [
        ("CC", CC_PATH, CC_PASSWORD),
        ("DL", DL_PATH, DL_PASSWORD),
    ]

    all_headers = None
    all_data = []
    source_labels = []

    for label, path, pwd in sources:
        print(f"\n  处理 {label}: {path.name}")
        dec_bytes = decrypt_file(path, pwd)

        for sheet_suffix in [" Employee list", " Employee list 离职"]:
            sheet_name = sheet_suffix
            # 尝试读取
            try:
                dec_bytes.seek(0)
                headers, rows = read_roster_sheet_from_bytes(
                    dec_bytes, sheet_name, label
                )
                print(f"    [{sheet_name}]: {len(rows)} 行, {len(headers)} 列")

                # 给数据行添加来源标记
                data_type = "离职" if "离职" in sheet_name else "在职"
                source_tag = f"{label}-{data_type}"

                if all_headers is None:
                    all_headers = headers
                else:
                    # 取较长的headers列表
                    if len(headers) > len(all_headers):
                        all_headers = headers

                for row in rows:
                    all_data.append(row)
                    source_labels.append(source_tag)

            except Exception as e:
                print(f"    [{sheet_name}]: 读取失败 - {e}")

    # 补齐列数（不同sheet可能列数不同）
    max_cols = len(all_headers)
    for i in range(len(all_data)):
        if len(all_data[i]) < max_cols:
            all_data[i] = all_data[i] + [None] * (max_cols - len(all_data[i]))
        elif len(all_data[i]) > max_cols:
            all_data[i] = all_data[i][:max_cols]

    # 添加"数据来源"列
    all_headers_with_source = all_headers + ["数据来源"]
    for i in range(len(all_data)):
        all_data[i].append(source_labels[i])

    # 删除空列col_1（第1列通常是空的）
    if all_headers_with_source and all_headers_with_source[0].startswith("col_"):
        all_headers_with_source = all_headers_with_source[1:]
        for i in range(len(all_data)):
            all_data[i] = all_data[i][1:]
        print(f"  已删除空列 col_1")

    print(f"\n  合计: {len(all_data)} 行, {len(all_headers_with_source)} 列")
    print(f"  列名: {all_headers_with_source}")

    return all_headers_with_source, all_data


# ==================== Step 2+3: [ERP系统接口]获取postLevel ====================
def step2_fetch_postlevel():
    """调用[ERP系统接口]获取岗级数据，返回 {personNo: postLevel} 字典"""
    print("\n" + "=" * 60)
    print("Step 2: 调用[ERP系统接口]获取岗级数据")
    print("=" * 60)

    auth_info = generate_auth_info()
    if not auth_info:
        print("  [错误] 无法生成认证信息")
        return {}

    payload = {"authInfo": auth_info}
    try:
        resp = requests.post(SAP_API_URL, json=payload, timeout=60)
        resp.raise_for_status()
        data = resp.json()

        # 调试：打印响应结构
        print(f"  响应类型: {type(data)}")
        if isinstance(data, dict):
            print(f"  响应keys: {list(data.keys())[:5]}")
            records = data.get("data", data)
            if isinstance(records, str):
                # data可能是JSON字符串
                try:
                    records = json.loads(records)
                except:
                    print(f"  [警告] data是字符串，尝试解析失败: {records[:200]}")
                    return {}
        else:
            records = data

        if isinstance(records, list) and records:
            print(f"  第一条记录: {records[0]}")

        level_map = {}
        for rec in records:
            if isinstance(rec, dict):
                pno = str(rec.get("personNo", "")).strip()
                level = str(rec.get("postLevel", "")).strip()
                if pno:
                    level_map[pno] = level
                    # 同时存入后6位作为key（兼容不同长度的员工号）
                    suffix = pno[-6:] if len(pno) >= 6 else pno
                    level_map[suffix] = level
            elif isinstance(rec, str):
                # 可能是逗号分隔的字符串
                print(f"  [警告] 跳过字符串记录: {rec[:100]}")

        print(f"  获取到 {len(level_map)} 条岗级记录")
        return level_map
    except Exception as e:
        print(f"  [错误] API调用失败: {e}")
        return {}


def step3_parse_postlevel(postlevel_str):
    """
    将postLevel字符串转换为大级别数值。

    规则：
    - "数字+字母"（如"7A"）→ 提取整数部分 → 7
    - "00" → 1
    - "06" → 6（先赋值，后续实习生检测可能覆盖为"-"）
    - "99" → "/"（派驻）
    - "字母+数字"（如"A1"）→ A=10, B=11, C=12 → 提取整数部分
    """
    if not postlevel_str or postlevel_str in ("", "None", "nan"):
        return None

    s = postlevel_str.strip().upper()

    # 特殊值
    if s == "00":
        return 1
    if s == "06":
        return 6
    if s == "99":
        return "/"

    # 字母+数字（如"A1", "B2", "C1"）
    if len(s) >= 2 and s[0].isalpha() and s[1:].isdigit():
        letter_map = {"A": 10, "B": 11, "C": 12}
        base = letter_map.get(s[0])
        if base is not None:
            return base

    # 数字+字母（如"7A", "8B"）→ 提取整数部分
    if s and s[0].isdigit():
        num_part = ""
        for ch in s:
            if ch.isdigit():
                num_part += ch
            else:
                break
        if num_part:
            return int(num_part)

    return None


# ==================== Step 4: 员工号匹配 ====================
def step4_match_roster(headers, data_rows, level_map):
    """
    按员工号后6位匹配SAP的postLevel，写入大级别。
    同时做实习生覆盖和员工号L开头处理。
    """
    print("\n" + "=" * 60)
    print("Step 3-4: 匹配大级别")
    print("=" * 60)

    # 找到列索引
    col_map = {h: i for i, h in enumerate(headers)}
    eid_col = col_map.get("员工号", -1)
    name_col = col_map.get("姓名", -1)
    emp_type_col = col_map.get("编制类型", -1)  # 编制类型
    category_col = col_map.get("人员编制", -1)  # 人员编制 (Salary/Hourly)
    position_col = col_map.get("岗位", -1)
    nature_col = col_map.get("员工性质", -1)
    contract_col = col_map.get("合同类型", -1)

    # 添加"大级别"列（在现有列后追加）
    if "大级别" not in headers:
        headers.append("大级别")
        level_col = len(headers) - 1
    else:
        level_col = headers.index("大级别")

    # 添加"蓝领白领"列
    if "蓝领白领" not in headers:
        headers.append("蓝领白领")
        bl_col = len(headers) - 1
    else:
        bl_col = headers.index("蓝领白领")

    matched_count = 0
    unmatched_count = 0
    l_prefix_count = 0

    for i, row in enumerate(data_rows):
        # 确保行足够长
        while len(row) < len(headers):
            row.append(None)

        eid = (
            str(row[eid_col]).strip()
            if eid_col >= 0 and row[eid_col] is not None
            else ""
        )
        position = (
            str(row[position_col]).strip()
            if position_col >= 0 and row[position_col] is not None
            else ""
        )
        emp_nature = (
            str(row[nature_col]).strip()
            if nature_col >= 0 and row[nature_col] is not None
            else ""
        )
        contract = (
            str(row[contract_col]).strip()
            if contract_col >= 0 and row[contract_col] is not None
            else ""
        )

        # 员工号以"L"开头 → 大级别=1（外包）
        if eid.startswith("L") or eid.startswith("l"):
            row[level_col] = 1
            l_prefix_count += 1
            continue

        # 取员工号后6位
        eid_suffix = eid[-6:] if len(eid) >= 6 else eid

        # 从SAP level_map匹配
        sap_level = level_map.get(eid_suffix)
        if sap_level:
            parsed = step3_parse_postlevel(sap_level)
            row[level_col] = parsed
            matched_count += 1
        else:
            # 未匹配：根据岗位判断
            if "总经理" in position:
                row[level_col] = "//"
            else:
                row[level_col] = "/"
            unmatched_count += 1

        # 实习生覆盖（合同类型="实习生" 或 员工性质="实习生" → 大级别="-"）
        if contract == "实习生" or emp_nature == "实习生":
            row[level_col] = "-"

    print(f"  SAP匹配成功: {matched_count}")
    print(f"  SAP未匹配: {unmatched_count}")
    print(f"  L开头外包: {l_prefix_count}")

    return headers, data_rows


# ==================== Step 5: 部门名称规范化 ====================
def step5_normalize_departments(headers, data_rows):
    """VBA中的15条部门规范化规则"""
    print("\n" + "=" * 60)
    print("Step 5: 部门名称规范化")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    company_col = col_map.get("公司", -1)
    name_col = col_map.get("姓名", -1)
    dept_col = col_map.get("部门", -1)
    dept2_col = col_map.get("二级部门", -1)

    changes = 0

    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        company = (
            str(row[company_col]).strip()
            if company_col >= 0 and row[company_col] is not None
            else ""
        )
        name = (
            str(row[name_col]).strip()
            if name_col >= 0 and row[name_col] is not None
            else ""
        )
        dept = (
            str(row[dept_col]).strip()
            if dept_col >= 0 and row[dept_col] is not None
            else ""
        )
        dept2 = (
            str(row[dept2_col]).strip()
            if dept2_col >= 0 and row[dept2_col] is not None
            else ""
        )

        original_dept = dept

        # 规则1: 管理部/总经办
        if dept in ("管理部", "总经办"):
            if "公共关系" in dept2:
                dept = "总经办-MPR"
            elif company == "[关联公司]":
                dept = "总经办-EM"
            else:
                dept = "总经办"

        # 规则2: 财务管理部 + [关联公司]
        if dept == "财务管理部" and company == "[关联公司]":
            dept = "财务管理部-EM"

        # 规则3: 采购部 + [关联公司]
        if dept == "采购部" and company == "[关联公司]":
            dept = "采购部-EM"

        # 规则4: 钟彬特殊
        if name == "钟彬":
            dept = "卓越运营部(大连)"

        # 规则5: 系统部
        if dept == "系统部":
            if "人机界面" in dept2:
                dept = "系统部-UI"
            else:
                dept = "系统部-系统"

        # 规则6: 硬件研发部
        if dept == "硬件研发部":
            if "共享技术" in dept2:
                dept = "硬件研发部-共享技术"
            elif "机械" in dept2:
                dept = "硬件研发部-机械"
            else:
                dept = "硬件研发部-电子"

        # 规则7: 质量管理部
        if dept == "质量管理部":
            if "客户" in dept2:
                dept = "质量管理部-客户质量"
            elif "实验" in dept2:
                dept = "质量管理部-实验室"
            elif "制造" in dept2:
                dept = "质量管理部-制造质量"
            else:
                dept = "质量管理部-经营质量"

        # 规则8: 大客户管理部
        if dept == "大客户管理部":
            if "代工" in dept2:
                dept = "大客户管理部-代工"
            elif "日系" in dept2:
                dept = "大客户管理部-日系"
            else:
                dept = "大客户管理部-自主"

        # 规则9: 业务策划管理部
        if dept == "业务策划管理部":
            if "产品策划日系" in dept2 or "产品日系" in dept2:
                dept = "业务策划管理部-产品-日系"
            elif "产品策划自主" in dept2 or "产品自主" in dept2:
                dept = "业务策划管理部-产品-自主"
            elif "产品策划代工" in dept2 or "产品代工" in dept2:
                dept = "业务策划管理部-产品-代工"
            elif "项目管理日系" in dept2 or "项目日系" in dept2:
                dept = "业务策划管理部-项目-日系"
            elif "项目管理自主" in dept2 or "项目自主" in dept2:
                dept = "业务策划管理部-项目-自主"
            else:
                dept = "业务策划管理部-项目-代工"

        if dept != original_dept:
            changes += 1
        row[dept_col] = dept

    print(f"  部门名称规范化: {changes} 条记录被修改")
    return headers, data_rows


# ==================== Step 6: 中心补全 ====================
def step6_fill_center(headers, data_rows):
    """空中心根据部门推断"""
    print("\n" + "=" * 60)
    print("Step 6: 中心补全")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    center_col = col_map.get("中心", -1)
    dept_col = col_map.get("部门", -1)

    filled = 0
    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        center = (
            str(row[center_col]).strip()
            if center_col >= 0 and row[center_col] is not None
            else ""
        )
        dept = (
            str(row[dept_col]).strip()
            if dept_col >= 0 and row[dept_col] is not None
            else ""
        )

        if center == "" or center == "None":
            if "-EM" in dept:
                row[center_col] = "益劢职能中心"
            elif dept in ("表面贴装技术部", "预装部", "工程技术部"):
                row[center_col] = "制造二中心"
            elif "-日系" in dept:
                row[center_col] = "日系业务中心"
            elif "-自主" in dept:
                row[center_col] = "自主业务中心"
            elif "-代工" in dept:
                row[center_col] = "代工业务中心"
            else:
                row[center_col] = "职能中心"
            filled += 1

    print(f"  补全中心: {filled} 条记录")
    return headers, data_rows


# ==================== Step 7: 衍生字段 ====================
def is_admin_dept(dept):
    """判断是否属于职能部门（专业管理）"""
    admin_keywords = [
        "行政管理部",
        "财务管理部",
        "信息技术部",
        "人力资源管理部",
        "采购部",
        "总经办",
    ]
    for kw in admin_keywords:
        if dept.startswith(kw):
            return True
    return False


def is_direct_dept(dept):
    """判断是否属于直接生产部门"""
    return dept in ("总装部", "预装部", "表面贴装技术部")


def step7_derive_fields(headers, data_rows):
    """从大级别衍生：整理后的岗位、某零部件类别、直接间接"""
    print("\n" + "=" * 60)
    print("Step 7: 衍生字段（整理后的岗位/某零部件类别/直接间接）")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    level_col = col_map.get("大级别", -1)
    dept_col = col_map.get("部门", -1)
    position_col = col_map.get("岗位", -1)
    category_col = col_map.get("人员编制", -1)
    nature_col = col_map.get("员工性质", -1)

    # 确保新列存在
    for new_col in ["整理后的岗位", "某零部件类别", "直接间接"]:
        if new_col not in headers:
            headers.append(new_col)

    position2_col = headers.index("整理后的岗位")
    fuao_col = headers.index("某零部件类别")
    direct_col = headers.index("直接间接")

    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        level = row[level_col]
        dept = str(row[dept_col]).strip() if row[dept_col] is not None else ""
        position = (
            str(row[position_col]).strip() if row[position_col] is not None else ""
        )
        category = (
            str(row[category_col]).strip() if row[category_col] is not None else ""
        )
        nature = str(row[nature_col]).strip() if row[nature_col] is not None else ""

        admin = is_admin_dept(dept)
        direct = is_direct_dept(dept)

        # 默认值
        row[position2_col] = "实习生"
        row[fuao_col] = "辅助生产-合同制"
        row[direct_col] = "间接"

        if level == 1:
            row[position2_col] = "外包"
            row[fuao_col] = "基本生产-劳务"
            row[direct_col] = "直接" if direct else "间接"

        elif level in (2, 3):
            row[position2_col] = "操作工"
            if "派遣" in nature:
                row[fuao_col] = "基本生产-劳务"
            else:
                row[fuao_col] = "基本生产-合同制"
            row[direct_col] = "直接" if direct else "间接"

        elif level in (4, 5):
            if category == "Indirect Salary":
                row[position2_col] = "现场/技术员/文员等"
                row[direct_col] = "间接"
            else:
                row[position2_col] = "班长/技术员"
                row[direct_col] = "直接" if direct else "间接"
            row[fuao_col] = "辅助生产-合同制"

        elif level == 6:
            row[position2_col] = "助理工程师"
            row[fuao_col] = (
                "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
            )
            row[direct_col] = "间接"

        elif level == 7:
            row[position2_col] = "工程师"
            row[fuao_col] = (
                "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
            )
            row[direct_col] = "间接"

        elif level == 8:
            if "学科经理" in position or "HRBP" in position or "主管" in position:
                row[position2_col] = "主管"
            else:
                row[position2_col] = "高级工程师"
            row[fuao_col] = (
                "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
            )
            row[direct_col] = "间接"

        elif level == 9:
            if "工程师" in position or "专家" in position:
                row[position2_col] = "首席工程师"
            else:
                row[position2_col] = "经理"
            row[fuao_col] = (
                "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
            )
            row[direct_col] = "间接"

        elif level == 10:
            if (
                "学科经理" in position
                or "部经理" in position
                or "部门经理" in position
                or "副经理" in position
            ):
                row[position2_col] = "高级经理"
                row[fuao_col] = "二级经理-专业管理-合同制"
            else:
                row[position2_col] = "高级首席工程师"
                row[fuao_col] = (
                    "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
                )
            row[direct_col] = "间接"

        elif level == 11:
            if "总监" in position:
                row[position2_col] = "总监"
                row[fuao_col] = "二级经理-专业管理-合同制"
            else:
                row[position2_col] = "总工程师"
                row[fuao_col] = (
                    "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
                )
            row[direct_col] = "间接"

        elif level == 12:
            if "总监" in position:
                row[position2_col] = "高级总监"
                row[fuao_col] = "二级经理-专业管理-合同制"
            else:
                row[position2_col] = "科学家"
                row[fuao_col] = (
                    "一般干部-专业管理-合同制" if admin else "一般干部-工程技术-合同制"
                )
            row[direct_col] = "间接"

        elif level == "/":
            row[position2_col] = "派驻"
            row[fuao_col] = "二级经理-专业管理-合同制"
            row[direct_col] = "间接"

        elif level == "//":
            row[position2_col] = "派驻-总经理"
            row[fuao_col] = "高级经理"
            row[direct_col] = "间接"

        elif level == "-":
            row[position2_col] = "实习生"
            row[fuao_col] = "辅助生产-合同制"
            row[direct_col] = "间接"

    print(f"  衍生字段已计算")
    return headers, data_rows


# ==================== Step 8: 蓝领白领 ====================
def step8_classify_collar(headers, data_rows):
    """蓝领白领分类。
    实习生判断：合同类型="实习生" 或 员工性质="实习生"
    """
    print("\n" + "=" * 60)
    print("Step 8: 蓝领白领分类")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    category_col = col_map.get("人员编制", -1)
    nature_col = col_map.get("员工性质", -1)
    contract_col = col_map.get("合同类型", -1)
    bl_col = col_map.get("蓝领白领", -1)

    if bl_col < 0:
        headers.append("蓝领白领")
        bl_col = len(headers) - 1

    counts = {"白领": 0, "蓝领": 0, "实习生": 0}

    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        category = (
            str(row[category_col]).strip()
            if category_col >= 0 and row[category_col] is not None
            else ""
        )
        nature = (
            str(row[nature_col]).strip()
            if nature_col >= 0 and row[nature_col] is not None
            else ""
        )
        contract = (
            str(row[contract_col]).strip()
            if contract_col >= 0 and row[contract_col] is not None
            else ""
        )

        if contract == "实习生" or nature == "实习生":
            row[bl_col] = "实习生"
            counts["实习生"] += 1
        elif category == "Salary":
            row[bl_col] = "白领"
            counts["白领"] += 1
        else:
            row[bl_col] = "蓝领"
            counts["蓝领"] += 1

    print(
        f"  白领: {counts['白领']}, 蓝领: {counts['蓝领']}, 实习生: {counts['实习生']}"
    )
    return headers, data_rows


# ==================== Step 9: 特殊修正 ====================
def step9_special_fixes(headers, data_rows):
    """特殊人物修正"""
    print("\n" + "=" * 60)
    print("Step 9: 特殊修正")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    name_col = col_map.get("姓名", -1)
    eid_col = col_map.get("员工号", -1)
    center_col = col_map.get("中心", -1)
    dept_col = col_map.get("部门", -1)

    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        name = (
            str(row[name_col]).strip()
            if name_col >= 0 and row[name_col] is not None
            else ""
        )
        eid = (
            str(row[eid_col]).strip()
            if eid_col >= 0 and row[eid_col] is not None
            else ""
        )

        if name == "王鑫" and eid == "110136":
            row[center_col] = "德系业务中心"
            row[dept_col] = "德系大客户管理部"
            print(f"  修正: 王鑫(110136) → 中心=德系业务中心, 部门=德系大客户管理部")

        if name == "王泽明" and eid == "120003":
            row[center_col] = "职能中心"
            print(f"  修正: 王泽明(120003) → 中心=职能中心")

    return headers, data_rows


# ==================== Step 10: 删除高管 ====================
def step10_remove_executives(headers, data_rows):
    """删除岗位为总经理/副总经理的员工"""
    print("\n" + "=" * 60)
    print("Step 10: 删除高管（总经理/副总经理）")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    position_col = col_map.get("岗位", -1)
    name_col = col_map.get("姓名", -1)

    before = len(data_rows)
    removed = []
    filtered = []

    for row in data_rows:
        position = (
            str(row[position_col]).strip()
            if position_col >= 0 and row[position_col] is not None
            else ""
        )
        name = (
            str(row[name_col]).strip()
            if name_col >= 0 and row[name_col] is not None
            else ""
        )

        if position in ("总经理", "副总经理"):
            removed.append(f"{name}({position})")
        else:
            filtered.append(row)

    print(f"  删除 {len(removed)} 人: {', '.join(removed)}")
    print(f"  剩余 {len(filtered)} 人")

    return headers, filtered


# ==================== Step 11: 大级别"/"的岗级识别 ====================
def step11_reclassify_slash(headers, data_rows):
    """大级别为"/"的人，根据岗位识别岗级"""
    print("\n" + "=" * 60)
    print("Step 11: 重新识别大级别为/的岗级")
    print("=" * 60)

    col_map = {h: i for i, h in enumerate(headers)}
    level_col = col_map.get("大级别", -1)
    position_col = col_map.get("岗位", -1)
    name_col = col_map.get("姓名", -1)

    reclassified = {"11": 0, "10": 0, "9": 0}

    for row in data_rows:
        while len(row) < len(headers):
            row.append(None)

        level = row[level_col]
        position = (
            str(row[position_col]).strip()
            if position_col >= 0 and row[position_col] is not None
            else ""
        )
        name = (
            str(row[name_col]).strip()
            if name_col >= 0 and row[name_col] is not None
            else ""
        )

        if level == "/":
            if "总监" in position:
                row[level_col] = 11
                reclassified["11"] += 1
            elif "部经理" in position or "部门经理" in position:
                row[level_col] = 10
                reclassified["10"] += 1
            else:
                row[level_col] = 9
                reclassified["9"] += 1

    print(
        f"  →11级: {reclassified['11']}, →10级: {reclassified['10']}, →9级: {reclassified['9']}"
    )
    return headers, data_rows


# ==================== Step 11.5: 替换中文括号 ====================
def step11_5_fix_parentheses(headers, data_rows):
    """将所有单元格中的中文括号（）替换为英文括号()"""
    print("\n" + "=" * 60)
    print("Step 11.5: 替换中文括号为英文括号")
    print("=" * 60)

    # 替换表头
    for i in range(len(headers)):
        if isinstance(headers[i], str):
            headers[i] = headers[i].replace("（", "(").replace("）", ")")

    # 替换数据
    count = 0
    for row in data_rows:
        for j in range(len(row)):
            val = row[j]
            if isinstance(val, str) and ("（" in val or "）" in val):
                row[j] = val.replace("（", "(").replace("）", ")")
                count += 1

    print(f"  替换了 {count} 个单元格中的中文括号")
    return headers, data_rows


# ==================== Step 12: 写入Excel ====================
def step12_write_excel(headers, data_rows):
    """写入Excel文件。
    直接写入 EXISTING_PATH（databases/员工花名册.xlsx）的"花名册" sheet，
    保留该文件中已有的"人数预算" sheet 不变。
    """
    print("\n" + "=" * 60)
    print(f"Step 12: 写入Excel")
    print("=" * 60)

    # === 2. 写入 EXISTING_PATH（覆盖花名册 sheet，保留人数预算 sheet）===
    if EXISTING_PATH.exists():
        try:
            wb = openpyxl.load_workbook(str(EXISTING_PATH))
        except Exception as e:
            print(f"  [错误] 无法打开 {EXISTING_PATH}: {e}")
            return
    else:
        wb = openpyxl.Workbook()

    # 如果"花名册" sheet已存在，删除后重建
    if "花名册" in wb.sheetnames:
        del wb["花名册"]

    ws = wb.create_sheet("花名册", 0)  # 插入到第一个位置

    # 写入标题行
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据行
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            if col_idx <= len(headers):
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(str(EXISTING_PATH))
    print(f"  已覆盖: {EXISTING_PATH}")
    print(f"  花名册: {len(data_rows)} 行, {len(headers)} 列")

    # 确认保留的sheet
    wb_check = openpyxl.load_workbook(str(EXISTING_PATH), read_only=True)
    print(f"  保留的sheets: {wb_check.sheetnames}")
    wb_check.close()


# ==================== 主流程 ====================
def main():
    print("花名册自动刷新")
    print()

    # Step 1: 解密合并
    headers, data_rows = step1_stack_sheets()

    # Step 2: 获取SAP岗级
    level_map = step2_fetch_postlevel()

    # Step 3-4: 匹配大级别
    headers, data_rows = step4_match_roster(headers, data_rows, level_map)

    # Step 5: 部门规范化
    headers, data_rows = step5_normalize_departments(headers, data_rows)

    # Step 6: 中心补全
    headers, data_rows = step6_fill_center(headers, data_rows)

    # Step 7: 衍生字段
    headers, data_rows = step7_derive_fields(headers, data_rows)

    # Step 8: 蓝领白领
    headers, data_rows = step8_classify_collar(headers, data_rows)

    # Step 9: 特殊修正
    headers, data_rows = step9_special_fixes(headers, data_rows)

    # Step 10: 删除高管
    headers, data_rows = step10_remove_executives(headers, data_rows)

    # Step 11: 重新识别/级
    headers, data_rows = step11_reclassify_slash(headers, data_rows)

    # Step 11.5: 替换中文括号为英文括号
    headers, data_rows = step11_5_fix_parentheses(headers, data_rows)

    # Step 12: 写入Excel
    step12_write_excel(headers, data_rows)

    print("\n" + "=" * 60)
    print("完成！")
    print("=" * 60)
    print()
    print("提示：花名册索引已更新。如需只重建花名册的向量索引（约50秒），运行：")
    print("  python -m src.api.server --update-db 员工花名册.xlsx")
    print("如需全量重建所有数据库索引（约10分钟），运行：")
    print("  python -m src.api.server --update")


if __name__ == "__main__":
    main()
