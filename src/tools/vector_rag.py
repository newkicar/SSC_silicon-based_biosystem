"""
向量化RAG知识库 —— 基于Embedding的语义检索

技术方案：
- 优先使用bge-m3 Embedding模型（通过Ollama OpenAI兼容API调用）
- 自动降级为TF-IDF（如果Embedding服务不可用）
- 自动将RAG_files/目录下的文档切片入库
- 余弦相似度匹配返回Top-K最相关段落
"""

import os
import re
import json
import pickle
import hashlib
import time
from pathlib import Path

import numpy as np
import requests as http_requests

from src.config.settings import RAG_DIR, DATA_DIR, DB_DIR

# ==================== Embedding配置 ====================
EMBEDDING_API_URL = "EMBEDDING_API_URL"
EMBEDDING_MODEL = "EMBEDDING_MODEL_NAME"
EMBEDDING_API_KEY = "EMBEDDING_API_KEY"
EMBEDDING_DIM = 1024
EMBEDDING_BATCH_SIZE = 32  # 每批发送的文本数量

# 索引持久化路径
INDEX_PATH = DATA_DIR / "rag_index.pkl"

# 全局单例
_index_data = None

# Embedding服务可用性标记
_embedding_available = None


def _check_embedding_available() -> bool:
    """检测Embedding服务是否可用"""
    global _embedding_available
    if _embedding_available is not None:
        return _embedding_available

    try:
        resp = http_requests.post(
            EMBEDDING_API_URL,
            json={"model": EMBEDDING_MODEL, "input": "测试"},
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {EMBEDDING_API_KEY}",
            },
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            dim = len(data["data"][0]["embedding"])
            _embedding_available = True
            print(f"[向量RAG] Embedding服务可用：{EMBEDDING_MODEL}（维度: {dim}）")
            return True
        else:
            _embedding_available = False
            print(f"[向量RAG] Embedding服务返回错误: {resp.status_code}，降级为TF-IDF")
            return False
    except Exception as e:
        _embedding_available = False
        print(f"[向量RAG] Embedding服务不可用: {e}，降级为TF-IDF")
        return False


def _get_embeddings(texts: list[str]) -> np.ndarray | None:
    """
    批量获取文本的Embedding向量。
    返回 shape=(len(texts), EMBEDDING_DIM) 的numpy数组。
    """
    all_embeddings = []

    for i in range(0, len(texts), EMBEDDING_BATCH_SIZE):
        batch = texts[i : i + EMBEDDING_BATCH_SIZE]
        try:
            resp = http_requests.post(
                EMBEDDING_API_URL,
                json={"model": EMBEDDING_MODEL, "input": batch},
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {EMBEDDING_API_KEY}",
                },
                timeout=60,
            )
            if resp.status_code != 200:
                print(f"[向量RAG] Embedding API错误: {resp.status_code}")
                return None

            data = resp.json()
            batch_embeddings = [item["embedding"] for item in data["data"]]
            all_embeddings.extend(batch_embeddings)
        except Exception as e:
            print(f"[向量RAG] Embedding请求失败: {e}")
            return None

    return np.array(all_embeddings, dtype=np.float32)


def _get_single_embedding(text: str) -> np.ndarray | None:
    """获取单个文本的Embedding向量"""
    result = _get_embeddings([text])
    if result is not None and len(result) > 0:
        return result[0]
    return None


def _cosine_similarity_batch(query_vec: np.ndarray, matrix: np.ndarray) -> np.ndarray:
    """计算查询向量与矩阵中每个向量的余弦相似度"""
    query_norm = np.linalg.norm(query_vec)
    if query_norm == 0:
        return np.zeros(matrix.shape[0])

    # 计算矩阵每行的范数
    matrix_norms = np.linalg.norm(matrix, axis=1)
    # 避免除零
    matrix_norms = np.where(matrix_norms == 0, 1, matrix_norms)

    similarities = np.dot(matrix, query_vec) / (matrix_norms * query_norm)
    return similarities


def _load_index():
    """加载或构建索引"""
    global _index_data
    if _index_data is not None:
        return _index_data

    if INDEX_PATH.exists():
        try:
            with open(INDEX_PATH, "rb") as f:
                _index_data = pickle.load(f)
            # 检查索引类型是否与当前模式匹配
            use_embedding = _check_embedding_available()
            index_has_embedding = "embeddings" in _index_data

            if use_embedding and not index_has_embedding:
                print("[向量RAG] 当前索引为TF-IDF模式，需要重建为Embedding模式")
                build_index()
            elif not use_embedding and index_has_embedding:
                print("[向量RAG] 当前索引为Embedding模式，需要重建为TF-IDF模式")
                build_index()
            else:
                return _index_data
        except Exception:
            pass

    # 索引不存在或格式不匹配，自动构建
    build_index()
    return _index_data


def _chunk_text(text: str, chunk_size: int = 300) -> list[str]:
    """
    将文本按段落/句子切片。

    切片策略：
    1. 按段落（双换行）切分
    2. 过短的段落与相邻段落合并
    3. 过长的段落按chunk_size二次切分
    """
    paragraphs = re.split(r"\n\s*\n", text.strip())

    chunks = []
    current_chunk = ""

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue

        if len(current_chunk) + len(para) + 1 <= chunk_size:
            current_chunk = f"{current_chunk}\n{para}" if current_chunk else para
        else:
            if current_chunk:
                chunks.append(current_chunk.strip())

            if len(para) > chunk_size:
                sentences = re.split(r"[。！？；\n]", para)
                current_chunk = ""
                for sent in sentences:
                    sent = sent.strip()
                    if not sent:
                        continue
                    if len(current_chunk) + len(sent) + 1 <= chunk_size:
                        current_chunk = (
                            f"{current_chunk}。{sent}" if current_chunk else sent
                        )
                    else:
                        if current_chunk:
                            chunks.append(current_chunk.strip())
                        current_chunk = sent
            else:
                current_chunk = para

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return chunks


def build_index():
    """
    构建向量索引。
    优先使用Embedding，不可用时降级为TF-IDF。
    """
    global _index_data

    txt_files = list(RAG_DIR.glob("*.txt"))
    if not txt_files:
        print("[向量RAG] RAG_files/目录下无.txt文件")
        _index_data = {"documents": [], "sources": [], "mode": "empty"}
        return 0

    documents = []
    sources = []

    for filepath in txt_files:
        filename = filepath.stem
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            print(f"[向量RAG] 读取{filepath.name}失败: {e}")
            continue

        chunks = _chunk_text(content)

        for chunk in chunks:
            documents.append(chunk)
            sources.append(filename)

    if not documents:
        _index_data = {"documents": [], "sources": [], "mode": "empty"}
        return 0

    use_embedding = _check_embedding_available()

    if use_embedding:
        # ========== Embedding模式 ==========
        print(f"[向量RAG] 使用Embedding模式构建索引（{len(documents)}个切片）...")
        start_time = time.time()

        embeddings = _get_embeddings(documents)

        if embeddings is not None:
            elapsed = time.time() - start_time
            _index_data = {
                "documents": documents,
                "sources": sources,
                "embeddings": embeddings,  # shape: (N, 1024)
                "mode": "embedding",
            }

            # 持久化索引
            INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(INDEX_PATH, "wb") as f:
                pickle.dump(_index_data, f)

            print(
                f"[向量RAG] Embedding索引构建完成：{len(txt_files)}个文件 → {len(documents)}个切片，耗时{elapsed:.1f}s"
            )
            return len(documents)
        else:
            print("[向量RAG] Embedding构建失败，降级为TF-IDF")
            # 继续走TF-IDF逻辑

    # ========== TF-IDF降级模式 ==========
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity as sklearn_cosine
    except ImportError:
        print("[向量RAG] sklearn未安装，无法使用TF-IDF降级模式")
        _index_data = {"documents": documents, "sources": sources, "mode": "keyword"}
        return len(documents)

    print(f"[向量RAG] 使用TF-IDF模式构建索引（{len(documents)}个切片）...")

    vectorizer = TfidfVectorizer(
        analyzer="char_wb",
        ngram_range=(2, 4),
        max_features=10000,
        sublinear_tf=True,
    )

    tfidf_matrix = vectorizer.fit_transform(documents)

    _index_data = {
        "documents": documents,
        "sources": sources,
        "vectorizer": vectorizer,
        "tfidf_matrix": tfidf_matrix,
        "mode": "tfidf",
    }

    # 持久化索引
    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(INDEX_PATH, "wb") as f:
        pickle.dump(_index_data, f)

    print(
        f"[向量RAG] TF-IDF索引构建完成：{len(txt_files)}个文件 → {len(documents)}个切片"
    )
    return len(documents)


def search(query: str, top_k: int = 5, min_score: float = 0.1) -> list[dict]:
    """
    语义检索。

    Args:
        query: 查询文本
        top_k: 返回最相关的K个结果
        min_score: 最低相似度阈值（0-1）

    Returns:
        list of {"content": str, "source": str, "score": float}
    """
    index = _load_index()

    if not index["documents"]:
        return []

    mode = index.get("mode", "tfidf")

    if mode == "embedding":
        # ========== Embedding检索 ==========
        query_vec = _get_single_embedding(query)
        if query_vec is None:
            # Embedding失败，尝试降级
            print("[向量RAG] 查询Embedding失败，尝试关键词匹配")
            return _fallback_keyword_search(query, index, top_k, min_score)

        similarities = _cosine_similarity_batch(query_vec, index["embeddings"])
        top_indices = similarities.argsort()[-top_k:][::-1]

        results = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= min_score:
                results.append(
                    {
                        "content": index["documents"][idx],
                        "source": index["sources"][idx],
                        "score": round(score, 3),
                    }
                )
        return results

    elif mode == "tfidf":
        # ========== TF-IDF检索 ==========
        from sklearn.metrics.pairwise import cosine_similarity as sklearn_cosine

        query_vec = index["vectorizer"].transform([query])
        similarities = sklearn_cosine(query_vec, index["tfidf_matrix"]).flatten()

        top_indices = similarities.argsort()[-top_k:][::-1]

        results = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= min_score:
                results.append(
                    {
                        "content": index["documents"][idx],
                        "source": index["sources"][idx],
                        "score": round(score, 3),
                    }
                )
        return results

    else:
        # ========== 关键词降级检索 ==========
        return _fallback_keyword_search(query, index, top_k, min_score)


def _fallback_keyword_search(
    query: str, index: dict, top_k: int, min_score: float
) -> list[dict]:
    """关键词降级检索（当Embedding和TF-IDF都不可用时）"""
    results = []
    keywords = re.findall(r"[\u4e00-\u9fff]{2,}", query)

    for i, doc in enumerate(index["documents"]):
        score = 0
        for kw in keywords:
            if kw in doc:
                score += 1
        if keywords:
            score = score / len(keywords)
        if score >= min_score:
            results.append(
                {
                    "content": doc,
                    "source": index["sources"][i],
                    "score": round(score, 3),
                }
            )

    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:top_k]


def search_formatted(query: str, top_k: int = 5) -> str:
    """
    语义检索，返回格式化的文本结果。
    与原有 _search_in_documents() 输出格式兼容。
    """
    results = search(query, top_k=top_k)

    if not results:
        return ""

    lines = []
    for r in results:
        lines.append(f"【{r['source']}】（相关度: {r['score']}）")
        lines.append(r["content"])
        lines.append("")

    return "\n".join(lines).strip()


def get_stats() -> dict:
    """获取索引统计信息"""
    index = _load_index()

    source_counts = {}
    for src in index["sources"]:
        source_counts[src] = source_counts.get(src, 0) + 1

    mode = index.get("mode", "unknown")

    return {
        "total_chunks": len(index["documents"]),
        "by_source": source_counts,
        "index_path": str(INDEX_PATH),
        "mode": mode,
        "embedding_available": _embedding_available,
    }


def rebuild_index():
    """强制重建RAG文档索引"""
    global _index_data
    _index_data = None
    if INDEX_PATH.exists():
        INDEX_PATH.unlink()
    return build_index()


def rebuild_db_index():
    """强制重建数据库索引（--update 模式使用）"""
    global _db_index_data
    _db_index_data = None
    if DB_INDEX_PATH.exists():
        DB_INDEX_PATH.unlink()
    return _build_db_index()


# ==================== 数据库Excel索引（通用，非硬编码） ====================
DB_INDEX_PATH = DATA_DIR / "db_index.pkl"
_db_index_data = None


def _excel_row_to_text(filepath: Path, headers: list, row: list) -> str:
    """
    将Excel一行数据转为自然语言文本块。
    通用逻辑：不硬编码任何字段名，自动遍历所有列。
    格式："来源:文件名\n字段1: 值1\n字段2: 值2\n..."
    这样AI可以通过语义匹配找到任何字段的值。
    """
    parts = [f"来源: {filepath.stem}"]
    for i, val in enumerate(row):
        if val is None or str(val).strip() == "" or str(val).strip().lower() == "none":
            continue
        header = headers[i] if i < len(headers) else f"列{i}"
        parts.append(f"{header}: {val}")
    return "\n".join(parts)


# 数据库Embedding专用batch_size（比RAG更大，减少API调用次数）
# 149184行 / 256 = 583次API调用，比之前4662次（batch_size=32）快8倍
DB_EMBEDDING_BATCH_SIZE = 256


def _get_db_embeddings(texts: list[str]) -> np.ndarray | None:
    """
    数据库专用批量Embedding：更大的batch + 进度日志。
    """
    all_embeddings = []
    total_batches = (
        len(texts) + DB_EMBEDDING_BATCH_SIZE - 1
    ) // DB_EMBEDDING_BATCH_SIZE

    for i in range(0, len(texts), DB_EMBEDDING_BATCH_SIZE):
        batch = texts[i : i + DB_EMBEDDING_BATCH_SIZE]
        batch_num = i // DB_EMBEDDING_BATCH_SIZE + 1
        # 每10个batch打印一次进度
        if batch_num % 10 == 1 or batch_num == total_batches:
            print(
                f"[数据库索引]   Embedding进度: {batch_num}/{total_batches} 批（{len(all_embeddings)}行已处理）"
            )
        try:
            resp = http_requests.post(
                EMBEDDING_API_URL,
                json={"model": EMBEDDING_MODEL, "input": batch},
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {EMBEDDING_API_KEY}",
                },
                timeout=120,  # 256条文本需要更长超时
            )
            if resp.status_code != 200:
                print(f"[数据库索引] Embedding API错误: {resp.status_code}")
                return None

            data = resp.json()
            batch_embeddings = [item["embedding"] for item in data["data"]]
            all_embeddings.extend(batch_embeddings)
        except Exception as e:
            print(f"[数据库索引] Embedding请求失败: {e}")
            return None

    return np.array(all_embeddings, dtype=np.float32)


def _build_db_index():
    """
    通用数据库索引构建。
    自动扫描 databases/ 下所有 Excel 文件，
    将每行数据的所有字段转为文本块后向量化。
    无论数据库有多少张表、多少个字段，都不需要修改代码。

    性能优化：
    - 使用更大的batch_size（256 vs 32），减少API调用次数
    - 添加进度日志，避免"卡住半小时"的误判
    - 启动时通过 _load_db_index() 调用，有缓存直接跳过
    - 幂等保护：已构建则跳过，防止重复构建
    """
    global _db_index_data

    # 幂等保护：如果已经在内存中构建过，直接返回
    if _db_index_data is not None and _db_index_data.get("documents"):
        return len(_db_index_data["documents"])

    if not DB_DIR.exists():
        print("[数据库索引] databases/ 目录不存在，跳过")
        _db_index_data = {"documents": [], "sources": [], "mode": "empty"}
        return 0

    xlsx_files = list(DB_DIR.glob("*.xlsx")) + list(DB_DIR.glob("*.xls"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]  # 排除临时文件
    if not xlsx_files:
        print("[数据库索引] databases/ 目录下无Excel文件")
        _db_index_data = {"documents": [], "sources": [], "mode": "empty"}
        return 0

    documents = []
    sources = []

    for filepath in xlsx_files:
        filename = filepath.stem
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # 第一行为表头
                headers = [
                    str(c).strip() if c else f"列{j}" for j, c in enumerate(rows[0])
                ]
                data_rows = rows[1:]
                for row in data_rows:
                    # 跳过全空行
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue
                    text = _excel_row_to_text(filepath, headers, list(row))
                    documents.append(text)
                    sources.append(f"{filename}/{ws_name}")
            wb.close()
        except Exception as e:
            print(f"[数据库索引] 读取{filepath.name}失败: {e}")
            continue

    if not documents:
        _db_index_data = {"documents": [], "sources": [], "mode": "empty"}
        return 0

    use_embedding = _check_embedding_available()

    if use_embedding:
        total_batches = (
            len(documents) + DB_EMBEDDING_BATCH_SIZE - 1
        ) // DB_EMBEDDING_BATCH_SIZE
        print(
            f"[数据库索引] 使用Embedding模式构建（{len(documents)}行数据，{total_batches}批，batch_size={DB_EMBEDDING_BATCH_SIZE}）..."
        )
        start_time = time.time()
        embeddings = _get_db_embeddings(documents)
        if embeddings is not None:
            elapsed = time.time() - start_time
            _db_index_data = {
                "documents": documents,
                "sources": sources,
                "embeddings": embeddings,
                "mode": "embedding",
            }
            DB_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(DB_INDEX_PATH, "wb") as f:
                pickle.dump(_db_index_data, f)
            print(
                f"[数据库索引] Embedding索引构建完成：{len(xlsx_files)}个文件 → {len(documents)}行，耗时{elapsed:.1f}s"
            )
            return len(documents)
        else:
            print("[数据库索引] Embedding构建失败，降级为TF-IDF")

    # TF-IDF降级
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
    except ImportError:
        print("[数据库索引] sklearn未安装，降级为关键词模式")
        _db_index_data = {"documents": documents, "sources": sources, "mode": "keyword"}
        return len(documents)

    print(f"[数据库索引] 使用TF-IDF模式构建（{len(documents)}行数据）...")
    vectorizer = TfidfVectorizer(
        analyzer="char_wb",
        ngram_range=(2, 4),
        max_features=20000,
        sublinear_tf=True,
    )
    tfidf_matrix = vectorizer.fit_transform(documents)
    _db_index_data = {
        "documents": documents,
        "sources": sources,
        "vectorizer": vectorizer,
        "tfidf_matrix": tfidf_matrix,
        "mode": "tfidf",
    }
    DB_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DB_INDEX_PATH, "wb") as f:
        pickle.dump(_db_index_data, f)
    print(
        f"[数据库索引] TF-IDF索引构建完成：{len(xlsx_files)}个文件 → {len(documents)}行"
    )
    return len(documents)


def rebuild_db_index_file(target_filename: str):
    """
    只重建指定文件的数据库索引（增量更新）。
    从已有缓存中删除该文件的旧记录，只对新文件做Embedding，然后合并。

    Args:
        target_filename: 文件名，如 "员工花名册.xlsx"
    """
    global _db_index_data

    target_path = DB_DIR / target_filename
    if not target_path.exists():
        print(f"[数据库索引] 文件不存在: {target_path}")
        return 0

    target_stem = target_path.stem

    # 1. 加载现有索引（如果有的话）
    existing_docs = []
    existing_sources = []
    existing_embeddings = None
    existing_mode = "empty"

    if DB_INDEX_PATH.exists():
        try:
            with open(DB_INDEX_PATH, "rb") as f:
                old_data = pickle.load(f)
            old_docs = old_data.get("documents", [])
            old_sources = old_data.get("sources", [])
            old_embeddings = old_data.get("embeddings")
            existing_mode = old_data.get("mode", "empty")

            # 2. 保留非目标文件的记录
            keep_indices = []
            for i, src in enumerate(old_sources):
                if not src.startswith(target_stem + "/"):
                    keep_indices.append(i)

            existing_docs = [old_docs[i] for i in keep_indices]
            existing_sources = [old_sources[i] for i in keep_indices]
            if old_embeddings is not None and len(old_embeddings) == len(old_docs):
                existing_embeddings = old_embeddings[keep_indices]

            print(
                f"[数据库索引] 保留 {len(existing_docs)} 条旧记录（已排除 {target_filename}）"
            )
        except Exception as e:
            print(f"[数据库索引] 加载旧索引失败: {e}")

    # 3. 只读取目标文件
    new_docs = []
    new_sources = []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(target_path, read_only=True, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c).strip() if c else f"列{j}" for j, c in enumerate(rows[0])]
            data_rows = rows[1:]
            for row in data_rows:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                text = _excel_row_to_text(target_path, headers, list(row))
                new_docs.append(text)
                new_sources.append(f"{target_stem}/{ws_name}")
        wb.close()
    except Exception as e:
        print(f"[数据库索引] 读取 {target_filename} 失败: {e}")
        return 0

    print(f"[数据库索引] {target_filename}: {len(new_docs)} 条新记录")

    # 4. 对新数据做Embedding
    new_embeddings = None
    use_embedding = _check_embedding_available()
    if use_embedding and new_docs:
        print(f"[数据库索引] 正在对 {target_filename} 做Embedding...")
        start_time = time.time()
        new_embeddings = _get_db_embeddings(new_docs)
        elapsed = time.time() - start_time
        if new_embeddings is not None:
            print(f"[数据库索引] {target_filename} Embedding完成，耗时{elapsed:.1f}s")
        else:
            print(f"[数据库索引] {target_filename} Embedding失败")

    # 5. 合并
    merged_docs = existing_docs + new_docs
    merged_sources = existing_sources + new_sources
    merged_mode = existing_mode

    if use_embedding and new_embeddings is not None:
        if existing_embeddings is not None:
            merged_embeddings = np.concatenate(
                [existing_embeddings, new_embeddings], axis=0
            )
        else:
            # 旧索引没有Embedding，需要对全部旧数据重新做
            print("[数据库索引] 旧索引无Embedding，重新对全量数据做Embedding...")
            merged_embeddings = _get_db_embeddings(merged_docs)
            if merged_embeddings is None:
                merged_embeddings = None
        merged_mode = "embedding"
    else:
        merged_embeddings = None
        if merged_docs and not use_embedding:
            # TF-IDF模式：需要全量重建
            try:
                from sklearn.feature_extraction.text import TfidfVectorizer

                vectorizer = TfidfVectorizer(
                    analyzer="char_wb",
                    ngram_range=(2, 4),
                    max_features=20000,
                    sublinear_tf=True,
                )
                tfidf_matrix = vectorizer.fit_transform(merged_docs)
                merged_mode = "tfidf"
                # 保存时用tfidf字段
                _db_index_data = {
                    "documents": merged_docs,
                    "sources": merged_sources,
                    "vectorizer": vectorizer,
                    "tfidf_matrix": tfidf_matrix,
                    "mode": merged_mode,
                }
                DB_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
                with open(DB_INDEX_PATH, "wb") as f:
                    pickle.dump(_db_index_data, f)
                print(f"[数据库索引] 增量更新完成（TF-IDF）：{len(merged_docs)} 条记录")
                return len(merged_docs)
            except ImportError:
                merged_mode = "keyword"

    # 保存Embedding索引
    save_data = {
        "documents": merged_docs,
        "sources": merged_sources,
        "mode": merged_mode,
    }
    if merged_embeddings is not None:
        save_data["embeddings"] = merged_embeddings

    DB_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DB_INDEX_PATH, "wb") as f:
        pickle.dump(save_data, f)

    # 更新内存缓存
    _db_index_data = save_data

    print(
        f"[数据库索引] 增量更新完成：{len(merged_docs)} 条记录（{target_filename}: +{len(new_docs)}）"
    )
    return len(merged_docs)


def _load_db_index():
    """加载或构建数据库索引"""
    global _db_index_data
    if _db_index_data is not None:
        return _db_index_data

    if DB_INDEX_PATH.exists():
        try:
            with open(DB_INDEX_PATH, "rb") as f:
                _db_index_data = pickle.load(f)
            use_embedding = _check_embedding_available()
            index_has_embedding = "embeddings" in _db_index_data
            if (use_embedding and not index_has_embedding) or (
                not use_embedding and index_has_embedding
            ):
                _build_db_index()
            else:
                return _db_index_data
        except Exception:
            pass

    _build_db_index()
    return _db_index_data


def search_db(query: str, top_k: int = 5, min_score: float = 0.1) -> list[dict]:
    """搜索数据库索引（databases/*.xlsx）"""
    index = _load_db_index()
    if not index["documents"]:
        return []

    mode = index.get("mode", "tfidf")

    if mode == "embedding":
        query_vec = _get_single_embedding(query)
        if query_vec is None:
            return _fallback_keyword_search(query, index, top_k, min_score)
        similarities = _cosine_similarity_batch(query_vec, index["embeddings"])
        top_indices = similarities.argsort()[-top_k:][::-1]
        results = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= min_score:
                results.append(
                    {
                        "content": index["documents"][idx],
                        "source": index["sources"][idx],
                        "score": round(score, 3),
                    }
                )
        return results

    elif mode == "tfidf":
        from sklearn.metrics.pairwise import cosine_similarity as sklearn_cosine

        query_vec = index["vectorizer"].transform([query])
        similarities = sklearn_cosine(query_vec, index["tfidf_matrix"]).flatten()
        top_indices = similarities.argsort()[-top_k:][::-1]
        results = []
        for idx in top_indices:
            score = float(similarities[idx])
            if score >= min_score:
                results.append(
                    {
                        "content": index["documents"][idx],
                        "source": index["sources"][idx],
                        "score": round(score, 3),
                    }
                )
        return results

    else:
        return _fallback_keyword_search(query, index, top_k, min_score)


def search_combined(
    query: str, top_k: int = 5, min_score: float = 0.1, max_total_chars: int = 8000
) -> str:
    """
    联合搜索：RAG政策文档 + databases/Excel数据。
    用于CLI端（员工终端），需要同时查政策和数据。

    Args:
        max_total_chars: 输出总字符数上限（防止token溢出，默认约2000 token）
    """
    results = []

    # 1. 搜索RAG政策文档
    rag_results = search(query, top_k=top_k, min_score=min_score)
    for r in rag_results:
        r["category"] = "政策文档"
        results.append(r)

    # 2. 搜索数据库Excel
    db_results = search_db(query, top_k=top_k, min_score=min_score)
    for r in db_results:
        r["category"] = "数据"
        results.append(r)

    # 3. 按相关度排序取Top
    results.sort(key=lambda x: x["score"], reverse=True)
    results = results[:top_k]

    if not results:
        return ""

    # 4. 格式化输出，带总字符数预算控制
    lines = []
    total_chars = 0
    # 每条结果最大字符数（根据结果数量动态调整）
    per_result_limit = min(600, max_total_chars // max(len(results), 1))

    for r in results:
        category = r.get("category", "")
        header = f"【{r['source']}】（{category}，相关度: {r['score']}）"
        content = r["content"][:per_result_limit]

        # 检查总预算
        entry_chars = len(header) + len(content) + 2
        if total_chars + entry_chars > max_total_chars:
            lines.append(
                f"...（已达到输出上限{max_total_chars}字符，剩余{len(results) - len(lines) // 3}条结果已省略）"
            )
            break

        lines.append(header)
        lines.append(content)
        lines.append("")
        total_chars += entry_chars

    return "\n".join(lines).strip()


# ==================== 兼容层 ====================
def vector_search_in_documents(keyword: str) -> str:
    """
    向量检索的兼容入口。
    可以直接替代原有 knowledge_base._search_in_documents() 的调用。
    """
    results = search(keyword, top_k=5, min_score=0.05)

    if not results:
        return f"未找到与'{keyword}'相关的内容。"

    lines = []
    for r in results:
        lines.append(f"【{r['source']}】（相关度: {r['score']}）")
        lines.append(r["content"][:500])
        lines.append("")

    return "\n".join(lines).strip()
